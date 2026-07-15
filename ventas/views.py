from __future__ import annotations

import csv
from datetime import date
import threading
import time
from io import BytesIO
from pathlib import Path
import uuid

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.core.files.storage import default_storage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .forms import CajaForm, CierreForm, JornadaForm, ProductoForm, RevisionDetallesForm
from .catalogo_imagenes import (
    adjuntar_imagenes_catalogo,
    adjuntar_imagenes_movimientos,
    adjuntar_imagenes_reporte,
    construir_tarjetas_presentacion,
)
from .models import Caja, Cierre, DetalleCierre, Jornada, Producto
from .services import (
    consolidado_de_jornada,
    consolidado_por_fecha,
    ejecutar_ocr,
    filas_detectadas_desde_ocr,
    clave_codigo,
    movimientos_de_jornada,
    parsear_detalles_texto,
    sugerir_detalles_desde_ocr,
    buscar_producto_por_codigo,
)


ORDEN_CAJAS = ['Presencial', 'Kio 1', 'Kio 2', 'Ventas online']


def cajas_operativas() -> list[Caja]:
    cajas = list(Caja.objects.filter(activa=True))
    orden = {nombre: posicion for posicion, nombre in enumerate(ORDEN_CAJAS)}
    return sorted(cajas, key=lambda caja: (orden.get(caja.nombre, 99), caja.nombre))


def interpretar_fecha(valor: str | None) -> date:
    try:
        return date.fromisoformat(valor or '')
    except ValueError:
        return timezone.localdate()


def _anotar_detecciones_catalogo(detecciones: list[dict]) -> tuple[list[dict], int]:
    """Marca filas OCR cuyo código no existe en el catálogo activo.

    La boleta puede traer códigos nuevos o una lectura OCR puede cortar mal un
    número. No se bloquea el guardado, pero la fila queda visible como
    REVISAR para que el usuario confirme antes de consolidar en Up Base.
    """
    if not detecciones:
        return [], 0
    productos = list(Producto.objects.filter(activo=True).only('nombre', 'codigo_vaso', 'codigo_botella'))
    anotadas: list[dict] = []
    desconocidos = 0
    for fila in detecciones:
        fila = dict(fila)
        codigo = str(fila.get('codigo') or '').strip()
        producto = buscar_producto_por_codigo(codigo, productos) if codigo else None
        fila['producto_catalogo'] = producto.nombre if producto else ''
        if codigo and producto is None:
            desconocidos += 1
            aviso = 'Código no registrado en el catálogo; revisar antes de guardar.'
            fila['alerta'] = f"{fila.get('alerta', '').strip()} {aviso}".strip()
            fila['codigo_no_registrado'] = True
        else:
            fila['codigo_no_registrado'] = False
        anotadas.append(fila)
    return anotadas, desconocidos


def dashboard(request):
    """Inicio agrupado por día; el detalle de métodos se consulta al abrir la fecha."""
    hoy = timezone.localdate()
    dias = list(
        Jornada.objects.filter(cierres__isnull=False)
        .values('fecha')
        .annotate(
            num_metodos=Count('id', distinct=True),
            num_cierres=Count('cierres', distinct=True),
        )
        .order_by('-fecha')[:12]
    )
    for dia in dias:
        dia['es_hoy'] = dia['fecha'] == hoy
    if not any(dia['es_hoy'] for dia in dias):
        dias.insert(0, {'fecha': hoy, 'num_metodos': 0, 'num_cierres': 0, 'es_hoy': True})

    contexto = {
        'dias': dias[:12],
        'total_productos': Producto.objects.filter(activo=True).count(),
        'pendientes_confirmar': Producto.objects.filter(activo=True, confirmado=False).count(),
        'total_cajas': Caja.objects.filter(activa=True).count(),
        'hoy': hoy,
    }
    return render(request, 'ventas/dashboard.html', contexto)


def cajas_lista(request):
    return render(request, 'ventas/cajas_lista.html', {'cajas': cajas_operativas()})


def caja_nueva(request):
    form = CajaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Caja registrada correctamente.')
        return redirect('cajas_lista')
    return render(request, 'ventas/formulario.html', {'form': form, 'titulo': 'Registrar caja', 'volver': reverse('cajas_lista')})


def productos_lista(request):
    q = request.GET.get('q', '').strip()
    vista = request.GET.get('vista', 'vaso').strip().lower()
    if vista not in {'general', 'vaso', 'botella'}:
        vista = 'vaso'
    productos = Producto.objects.filter(activo=True)
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | Q(codigo_vaso__icontains=q) | Q(codigo_botella__icontains=q)
            | Q(descripcion_vaso__icontains=q) | Q(descripcion_botella__icontains=q)
        )
    if vista == 'general':
        tarjetas, total_con_foto = adjuntar_imagenes_catalogo(productos)
    else:
        tarjetas, total_con_foto = construir_tarjetas_presentacion(productos, vista)
    return render(request, 'ventas/productos_lista.html', {
        'tarjetas': tarjetas,
        'vista': vista,
        'q': q,
        'total_con_foto': total_con_foto,
        'total_visibles': len(tarjetas),
    })


def producto_nuevo(request):
    form = ProductoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Producto registrado correctamente.')
        return redirect('productos_lista')
    return render(request, 'ventas/formulario.html', {'form': form, 'titulo': 'Nuevo producto', 'volver': reverse('productos_lista')})


def producto_editar(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Producto actualizado correctamente.')
        return redirect('productos_lista')
    return render(request, 'ventas/formulario.html', {
        'form': form,
        'titulo': f'Editar: {producto.nombre}',
        'volver': reverse('productos_lista'),
        'producto': producto,
    })


def _usos_de_producto(producto: Producto) -> tuple[int, int, bool]:
    """Cuenta usos del código y confirma si otra relación preserva la presentación botella."""
    claves = {
        clave_codigo(codigo)
        for codigo in (producto.codigo_vaso, producto.codigo_botella)
        if codigo
    }
    detalles_usados = sum(
        1 for codigo in DetalleCierre.objects.values_list('codigo_leido', flat=True).iterator()
        if clave_codigo(codigo) in claves
    )
    selecciones_conversion = producto.cierres_convertidos.count()

    respaldo_botella = False
    if producto.codigo_botella and not producto.codigo_vaso:
        misma_botella = clave_codigo(producto.codigo_botella)
        respaldo_botella = any(
            candidato.convertible and candidato.codigo_vaso
            and clave_codigo(candidato.codigo_botella) == misma_botella
            for candidato in Producto.objects.filter(activo=True).exclude(pk=producto.pk)
        )
    return detalles_usados, selecciones_conversion, respaldo_botella


def producto_eliminar(request, pk: int):
    producto = get_object_or_404(Producto, pk=pk)
    detalles_usados, selecciones_conversion, respaldo_botella = _usos_de_producto(producto)
    tiene_usos = bool(detalles_usados or selecciones_conversion)
    bloqueado = tiene_usos and not respaldo_botella

    if request.method == 'POST':
        if bloqueado:
            messages.warning(
                request,
                'No se eliminó el producto porque ya participa en cierres registrados. '
                'Así se evita modificar reportes históricos.'
            )
            return redirect('producto_editar', pk=producto.pk)
        nombre = producto.nombre
        producto.delete()
        messages.success(request, f'Producto eliminado: {nombre}.')
        return redirect(f"{reverse('productos_lista')}?vista=general")

    return render(request, 'ventas/producto_eliminar.html', {
        'producto': producto,
        'detalles_usados': detalles_usados,
        'selecciones_conversion': selecciones_conversion,
        'respaldo_botella': respaldo_botella,
        'bloqueado': bloqueado,
    })


def _eliminar_cargas_pendientes(lote: dict | None) -> None:
    """Retira imágenes subidas de un lote reemplazado antes de procesarse."""
    for slot in (lote or {}).get('slots', []):
        imagen = slot.get('imagen', '')
        if imagen:
            default_storage.delete(imagen)


def procesar_dia(request):
    """Recibe hasta cuatro fotografías y las prepara para una pantalla real de análisis."""
    cajas = cajas_operativas()
    fecha = interpretar_fecha(request.POST.get('fecha') if request.method == 'POST' else request.GET.get('fecha'))
    errores: list[str] = []

    if request.method == 'POST':
        slots: list[dict] = []
        for caja in cajas:
            imagen = request.FILES.get(f'imagen_{caja.pk}')
            if not imagen:
                continue
            modalidad = request.POST.get(f'modalidad_{caja.pk}', Cierre.Modalidad.VASO)
            if modalidad not in Cierre.Modalidad.values:
                modalidad = Cierre.Modalidad.VASO
            motivo = request.POST.get(f'motivo_{caja.pk}', '').strip()
            extension = Path(imagen.name).suffix.lower() or '.jpg'
            nombre_temporal = default_storage.save(
                f'cargas_pendientes/{uuid.uuid4().hex}{extension}', imagen
            )
            slots.append({
                'caja_id': caja.pk,
                'caja_nombre': caja.nombre,
                'modalidad': modalidad,
                'motivo': motivo,
                'imagen': nombre_temporal,
                'archivo': Path(imagen.name).name,
            })

        if not slots:
            errores.append('Seleccione por lo menos una imagen. Las casillas que no se utilicen pueden quedar vacías.')
        else:
            _eliminar_cargas_pendientes(request.session.pop('analisis_lote_pendiente', None))
            request.session['analisis_lote_pendiente'] = {
                'fecha': fecha.isoformat(),
                'slots': slots,
            }
            return redirect('procesando_lote')

    return render(request, 'ventas/procesar_dia.html', {
        'cajas': cajas,
        'fecha': fecha,
        'errores': errores,
        'modalidades': Cierre.Modalidad.choices,
    })


def procesando_lote(request):
    """Muestra la espera visual antes de ejecutar OCR, con las fotos ya recibidas."""
    lote = request.session.get('analisis_lote_pendiente')
    if not lote or not lote.get('slots'):
        return redirect('procesar_dia')
    fecha = interpretar_fecha(lote.get('fecha'))
    slots = []
    for slot in lote['slots']:
        slots.append({
            **slot,
            'imagen_url': default_storage.url(slot['imagen']),
        })
    return render(request, 'ventas/analizando_lote.html', {
        'fecha': fecha,
        'slots': slots,
        'cantidad_imagenes': len(slots),
    })


def ejecutar_analisis_lote(request):
    """Ejecuta OCR después de que la pantalla visible de procesamiento ya fue mostrada."""
    if request.method != 'POST':
        return redirect('procesar_dia')
    lote = request.session.get('analisis_lote_pendiente')
    if not lote or not lote.get('slots'):
        destino = reverse('revisar_lote') if request.session.get('cierres_lote_actual') else reverse('procesar_dia')
        return JsonResponse({'redirect': destino})

    fecha = interpretar_fecha(lote.get('fecha'))
    cierres_creados: list[int] = []
    try:
        with transaction.atomic():
            for slot in lote['slots']:
                caja = get_object_or_404(Caja, pk=slot['caja_id'], activa=True)
                jornada, _ = Jornada.objects.get_or_create(fecha=fecha, caja=caja)
                siguiente = (jornada.cierres.order_by('-numero').values_list('numero', flat=True).first() or 0) + 1
                cierre = Cierre.objects.create(
                    jornada=jornada,
                    numero=siguiente,
                    modalidad_periodo=slot['modalidad'],
                    motivo_cambio=slot['motivo'],
                    imagen=slot['imagen'],
                )
                texto, estado = ejecutar_ocr(cierre.imagen.path)
                cierre.ocr_texto = texto
                cierre.ocr_estado = estado
                cierre.save(update_fields=['ocr_texto', 'ocr_estado'])
                cierres_creados.append(cierre.pk)
    except Exception:
        return JsonResponse({
            'error': 'No se pudo completar el análisis. Las imágenes siguen disponibles para volver a intentarlo.'
        }, status=500)

    request.session.pop('analisis_lote_pendiente', None)
    request.session['cierres_lote_actual'] = cierres_creados
    request.session['fecha_lote_actual'] = fecha.isoformat()
    messages.info(
        request,
        f'Se analizaron {len(cierres_creados)} boleta(s). Revise las cantidades detectadas; las cajas vacías fueron omitidas.'
    )
    destino = reverse('revisar_lote')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'redirect': destino})
    return redirect(destino)



def _filas_editables_desde_texto(texto: str, detecciones: list[dict]) -> list[dict]:
    """Convierte el texto codigo,cantidad en filas para una tabla editable."""
    detecciones_por_codigo: dict[str, dict] = {}
    for fila in detecciones or []:
        codigo = str(fila.get('codigo') or '').strip()
        if codigo:
            detecciones_por_codigo.setdefault(clave_codigo(codigo), fila)

    filas, errores = parsear_detalles_texto(texto or '')
    filas_editables: list[dict] = []
    productos = list(Producto.objects.filter(activo=True).only('nombre', 'codigo_vaso', 'codigo_botella'))
    for indice, (codigo, cantidad) in enumerate(filas, start=1):
        deteccion = detecciones_por_codigo.get(clave_codigo(codigo), {})
        producto = buscar_producto_por_codigo(codigo, productos)
        alerta = str(deteccion.get('alerta') or '').strip()
        if producto is None:
            alerta = f"{alerta} Código no registrado en el catálogo; revisar antes de guardar.".strip()
        estado = 'Revisar' if alerta or cantidad < 0 else 'Detectado'
        filas_editables.append({
            'indice': indice,
            'codigo': codigo,
            'cantidad': cantidad,
            'descripcion': deteccion.get('descripcion', ''),
            'producto_catalogo': producto.nombre if producto else '',
            'estado': estado,
            'alerta': alerta,
            'no_registrado': producto is None,
        })

    if not filas_editables and (texto or '').strip():
        for indice, linea in enumerate((texto or '').splitlines(), start=1):
            if not linea.strip():
                continue
            filas_editables.append({
                'indice': indice,
                'codigo': '',
                'cantidad': '',
                'descripcion': linea.strip(),
                'producto_catalogo': '',
                'estado': 'Revisar',
                'alerta': 'No se pudo interpretar esta línea; complete código y cantidad.',
                'no_registrado': True,
            })
    return filas_editables


def _texto_desde_tabla_editable(request, cierre: Cierre) -> str:
    codigos = request.POST.getlist(f'codigo_{cierre.pk}[]')
    cantidades = request.POST.getlist(f'cantidad_{cierre.pk}[]')
    lineas: list[str] = []
    for codigo, cantidad in zip(codigos, cantidades):
        codigo = str(codigo or '').strip()
        cantidad = str(cantidad or '').strip()
        if not codigo and not cantidad:
            continue
        lineas.append(f'{codigo},{cantidad}')
    if lineas:
        return '\n'.join(lineas)
    return request.POST.get(f'detalles_{cierre.pk}', '').strip()

def revisar_lote(request):
    ids = request.session.get('cierres_lote_actual', [])
    fecha = interpretar_fecha(request.session.get('fecha_lote_actual'))
    cierres = {
        cierre.jornada.caja_id: cierre
        for cierre in Cierre.objects.filter(pk__in=ids).select_related('jornada', 'jornada__caja').prefetch_related('detalles')
    }
    slots = []
    for caja in cajas_operativas():
        cierre = cierres.get(caja.pk)
        inicial = ''
        if cierre:
            inicial = '\n'.join(f'{d.codigo_leido},{d.cantidad_acumulada}' for d in cierre.detalles.all())
            inicial = inicial or sugerir_detalles_desde_ocr(cierre.ocr_texto)
        detecciones_base = filas_detectadas_desde_ocr(cierre.ocr_texto) if cierre else []
        detecciones, codigos_no_registrados = _anotar_detecciones_catalogo(detecciones_base)
        filas_editables = _filas_editables_desde_texto(inicial, detecciones) if cierre else []
        slots.append({
            'caja': caja,
            'cierre': cierre,
            'inicial': inicial,
            'filas_editables': filas_editables,
            'detecciones': detecciones,
            'codigos_no_registrados': codigos_no_registrados,
            'errores': [],
        })

    if request.method == 'POST':
        procesados = 0
        pendientes = 0
        for slot in slots:
            cierre = slot['cierre']
            if not cierre:
                continue
            texto = _texto_desde_tabla_editable(request, cierre)
            slot['inicial'] = texto
            slot['filas_editables'] = _filas_editables_desde_texto(texto, slot.get('detecciones', []))
            if not texto:
                pendientes += 1
                continue
            filas, errores = parsear_detalles_texto(texto)
            if errores:
                slot['errores'] = errores
                pendientes += 1
                continue
            with transaction.atomic():
                cierre.detalles.all().delete()
                DetalleCierre.objects.bulk_create([
                    DetalleCierre(cierre=cierre, codigo_leido=codigo, cantidad_acumulada=cantidad)
                    for codigo, cantidad in filas
                ])
                cierre.confirmado = True
                cierre.save(update_fields=['confirmado'])
            procesados += 1

        if any(slot['errores'] for slot in slots):
            messages.warning(request, 'Algunas cajas tienen datos que corregir. Las cajas válidas ya pueden guardarse al volver a procesar.')
        elif procesados or pendientes:
            if procesados:
                messages.success(request, f'Se validaron {procesados} cierre(s).')
            if pendientes:
                messages.info(request, f'{pendientes} imagen(es) quedaron pendientes de digitación; no bloquean el consolidado.')
            return redirect('resumen_dia', fecha_iso=fecha.isoformat())

    return render(request, 'ventas/revisar_lote.html', {'slots': slots, 'fecha': fecha})



def _filas_para_escritura(fecha: date) -> list[dict]:
    """Consolidado ordenado para la escritura automática en Up Base."""
    consolidado, _ = consolidado_por_fecha(fecha)
    consolidado, _ = adjuntar_imagenes_reporte(consolidado)
    return consolidado


def _escribir_en_upbase_en_segundo_plano(filas: list[dict], espera: int, pausa: float) -> None:
    """Digita el consolidado en Up Base usando pegado rápido por portapapeles.

    Secuencia por fila: código → Enter → Arriba → Derecha → Derecha → cantidad → Enter.
    El uso del portapapeles es más rápido que escribir carácter por carácter.
    """
    try:
        import pyautogui
    except Exception:
        return
    try:
        import pyperclip
    except Exception:
        pyperclip = None

    def escribir_texto(valor: str) -> None:
        valor = str(valor)
        if pyperclip is not None:
            pyperclip.copy(valor)
            pyautogui.hotkey('ctrl', 'v')
        else:
            pyautogui.write(valor, interval=0.002)

    try:
        pyautogui.FAILSAFE = True
        pausa = max(0.03, float(pausa))
        pyautogui.PAUSE = min(0.04, pausa)
        time.sleep(max(1, int(espera)))
        for fila in filas:
            codigo = str(fila.get('codigo_final', '')).strip()
            cantidad = str(fila.get('cantidad', '')).strip()
            if not codigo or not cantidad:
                continue
            escribir_texto(codigo)
            pyautogui.press('enter')
            time.sleep(max(0.025, pausa * 0.45))
            pyautogui.press('up')
            pyautogui.press('right', presses=2, interval=max(0.005, pausa * 0.12))
            escribir_texto(cantidad)
            pyautogui.press('enter')
            time.sleep(pausa)
    except Exception:
        # Si el usuario mueve el mouse a una esquina o cambia de ventana en mal momento, se detiene sin romper Django.
        return

def ayuda_escritura_upbase(request, fecha_iso: str):
    fecha = interpretar_fecha(fecha_iso)
    consolidado = _filas_para_escritura(fecha)
    total_cantidad = sum(int(fila.get('cantidad') or 0) for fila in consolidado)
    pyautogui_disponible = True
    try:
        import pyautogui  # noqa: F401
    except Exception:
        pyautogui_disponible = False
    return render(request, 'ventas/ayuda_escritura_upbase.html', {
        'fecha': fecha,
        'consolidado': consolidado,
        'total_codigos': len(consolidado),
        'total_cantidad': total_cantidad,
        'pyautogui_disponible': pyautogui_disponible,
        'espera_default': 7,
        'pausa_default': '0.08',
    })


def iniciar_escritura_upbase(request, fecha_iso: str):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    try:
        import pyautogui  # noqa: F401
    except Exception:
        return JsonResponse({
            'error': 'Falta instalar pyautogui. Ejecute INSTALAR_AYUDA_ESCRITURA.bat y vuelva a abrir la aplicación.'
        }, status=500)

    fecha = interpretar_fecha(fecha_iso)
    filas = _filas_para_escritura(fecha)
    if not filas:
        return JsonResponse({'error': 'No hay códigos consolidados para escribir.'}, status=400)

    try:
        espera = int(request.POST.get('espera', '7'))
    except ValueError:
        espera = 7
    espera = min(max(espera, 3), 20)
    try:
        pausa = float(request.POST.get('pausa', '0.25'))
    except ValueError:
        pausa = 0.25
    pausa = min(max(pausa, 0.03), 1.50)

    hilo = threading.Thread(
        target=_escribir_en_upbase_en_segundo_plano,
        args=(filas, espera, pausa),
        daemon=True,
    )
    hilo.start()
    duracion_estimada = espera + max(2.0, len(filas) * (float(pausa) + 0.42))
    return JsonResponse({
        'ok': True,
        'espera': espera,
        'total': len(filas),
        'duracion_estimada': round(duracion_estimada, 2),
        'mensaje': 'Coloque el cursor en la primera celda de Código de Up Base. La escritura empezará al terminar el conteo.'
    })


def resumen_dia(request, fecha_iso: str):
    fecha = interpretar_fecha(fecha_iso)
    jornadas = list(
        Jornada.objects.filter(fecha=fecha).select_related('caja').prefetch_related('cierres').order_by('caja__nombre')
    )
    consolidado, advertencias = consolidado_por_fecha(fecha)
    consolidado, _ = adjuntar_imagenes_reporte(consolidado)
    return render(request, 'ventas/resumen_dia.html', {
        'fecha': fecha,
        'jornadas': jornadas,
        'total_metodos': len(jornadas),
        'total_cierres': sum(len(jornada.cierres.all()) for jornada in jornadas),
        'total_codigos': len(consolidado),
        'consolidado': consolidado,
        'advertencias': advertencias,
    })


def eliminar_registros_dia(request, fecha_iso: str):
    """Elimina todos los cierres de una fecha después de confirmación explícita."""
    try:
        fecha = date.fromisoformat(fecha_iso)
    except ValueError as exc:
        raise Http404('Fecha no válida.') from exc

    jornadas = list(
        Jornada.objects.filter(fecha=fecha)
        .select_related('caja')
        .prefetch_related('cierres')
        .order_by('caja__nombre')
    )
    cierres_qs = Cierre.objects.filter(jornada__fecha=fecha)
    total_cierres = cierres_qs.count()
    total_detalles = DetalleCierre.objects.filter(cierre__jornada__fecha=fecha).count()
    metodos = [
        {'nombre': jornada.caja.nombre, 'cierres': jornada.cierres.count()}
        for jornada in jornadas
    ]

    if request.method == 'POST':
        if not total_cierres and not jornadas:
            messages.info(request, 'No existen registros para eliminar en esa fecha.')
            return redirect('dashboard')

        imagenes_boleta = [
            nombre for nombre in cierres_qs.exclude(imagen='').values_list('imagen', flat=True) if nombre
        ]
        with transaction.atomic():
            Jornada.objects.filter(fecha=fecha).delete()

        for imagen in imagenes_boleta:
            default_storage.delete(imagen)

        if request.session.get('fecha_lote_actual') == fecha.isoformat():
            request.session.pop('cierres_lote_actual', None)
            request.session.pop('fecha_lote_actual', None)
        messages.success(
            request,
            f'Se eliminaron los registros del {fecha:%d/%m/%Y}: {total_cierres} cierre(s) y {total_detalles} fila(s) de productos.'
        )
        return redirect('dashboard')

    return render(request, 'ventas/dia_eliminar.html', {
        'fecha': fecha,
        'metodos': metodos,
        'total_cierres': total_cierres,
        'total_detalles': total_detalles,
        'tiene_registros': bool(jornadas or total_cierres),
        'es_hoy': fecha == timezone.localdate(),
    })


def jornada_nueva(request):
    form = JornadaForm(request.POST or None, initial={'fecha': timezone.localdate()})
    if request.method == 'POST' and form.is_valid():
        jornada = form.save()
        messages.success(request, 'Jornada creada. Ahora registre el primer cierre acumulado.')
        return redirect('jornada_detalle', pk=jornada.pk)
    return render(request, 'ventas/formulario.html', {'form': form, 'titulo': 'Nueva jornada manual', 'volver': reverse('dashboard')})


def jornada_detalle(request, pk: int):
    jornada = get_object_or_404(Jornada.objects.select_related('caja'), pk=pk)
    movimientos, advertencias = movimientos_de_jornada(jornada)
    movimientos, _ = adjuntar_imagenes_movimientos(movimientos)
    consolidado = consolidado_de_jornada(jornada)
    consolidado, _ = adjuntar_imagenes_reporte(consolidado)
    cierres = jornada.cierres.prefetch_related('detalles', 'productos_convertir').order_by('numero')
    return render(request, 'ventas/jornada_detalle.html', {
        'jornada': jornada,
        'cierres': cierres,
        'movimientos': movimientos,
        'advertencias': advertencias,
        'consolidado': consolidado,
    })


def cierre_nuevo(request, jornada_pk: int):
    jornada = get_object_or_404(Jornada, pk=jornada_pk)
    siguiente_numero = (jornada.cierres.order_by('-numero').values_list('numero', flat=True).first() or 0) + 1
    form = CierreForm(request.POST or None, request.FILES or None, initial={'numero': siguiente_numero})
    if request.method == 'POST' and form.is_valid():
        cierre = form.save(commit=False)
        cierre.jornada = jornada
        cierre.save()
        form.save_m2m()
        if cierre.imagen:
            texto, estado = ejecutar_ocr(cierre.imagen.path)
            cierre.ocr_texto = texto
            cierre.ocr_estado = estado
            cierre.save(update_fields=['ocr_texto', 'ocr_estado'])
        messages.info(request, 'Cierre creado. Revise y confirme las cantidades acumuladas de la boleta.')
        return redirect('cierre_revisar', pk=cierre.pk)
    return render(request, 'ventas/cierre_form.html', {'form': form, 'jornada': jornada, 'titulo': 'Registrar cierre acumulado'})


def cierre_editar(request, pk: int):
    cierre = get_object_or_404(Cierre, pk=pk)
    imagen_anterior = cierre.imagen.name if cierre.imagen else ''
    form = CierreForm(request.POST or None, request.FILES or None, instance=cierre)
    if request.method == 'POST' and form.is_valid():
        cierre = form.save()
        imagen_nueva = cierre.imagen.name if cierre.imagen else ''
        if cierre.imagen and imagen_nueva != imagen_anterior:
            texto, estado = ejecutar_ocr(cierre.imagen.path)
            cierre.ocr_texto = texto
            cierre.ocr_estado = estado
            cierre.save(update_fields=['ocr_texto', 'ocr_estado'])
        messages.success(request, 'Datos del cierre actualizados.')
        return redirect('cierre_revisar', pk=cierre.pk)
    return render(request, 'ventas/cierre_form.html', {'form': form, 'jornada': cierre.jornada, 'titulo': f'Editar cierre {cierre.numero}'})


def cierre_revisar(request, pk: int):
    cierre = get_object_or_404(Cierre.objects.select_related('jornada', 'jornada__caja'), pk=pk)
    detalles_existentes = '\n'.join(
        f'{detalle.codigo_leido},{detalle.cantidad_acumulada}' for detalle in cierre.detalles.all()
    )
    inicial = detalles_existentes or sugerir_detalles_desde_ocr(cierre.ocr_texto)
    form = RevisionDetallesForm(request.POST or None, initial={'detalles_texto': inicial})

    if request.method == 'POST' and form.is_valid():
        filas, errores = parsear_detalles_texto(form.cleaned_data['detalles_texto'])
        if errores:
            for error in errores:
                form.add_error('detalles_texto', error)
        else:
            with transaction.atomic():
                cierre.detalles.all().delete()
                DetalleCierre.objects.bulk_create([
                    DetalleCierre(cierre=cierre, codigo_leido=codigo, cantidad_acumulada=cantidad)
                    for codigo, cantidad in filas
                ])
                cierre.confirmado = True
                cierre.save(update_fields=['confirmado'])
            messages.success(request, 'Cierre validado. El consolidado para Up Base fue actualizado.')
            return redirect('jornada_detalle', pk=cierre.jornada_id)

    detecciones, codigos_no_registrados = _anotar_detecciones_catalogo(filas_detectadas_desde_ocr(cierre.ocr_texto))
    return render(request, 'ventas/cierre_revisar.html', {
        'cierre': cierre,
        'form': form,
        'detecciones': detecciones,
        'codigos_no_registrados': codigos_no_registrados,
    })


def exportar_csv(request, pk: int):
    jornada = get_object_or_404(Jornada.objects.select_related('caja'), pk=pk)
    return respuesta_csv(consolidado_de_jornada(jornada), f'upbase_{jornada.fecha}_{jornada.caja.nombre}.csv')


def exportar_xlsx(request, pk: int):
    jornada = get_object_or_404(Jornada.objects.select_related('caja'), pk=pk)
    return respuesta_xlsx(consolidado_de_jornada(jornada), f'upbase_{jornada.fecha}_{jornada.caja.nombre}.xlsx')


def exportar_dia_csv(request, fecha_iso: str):
    fecha = interpretar_fecha(fecha_iso)
    consolidado, _ = consolidado_por_fecha(fecha)
    return respuesta_csv(consolidado, f'upbase_consolidado_{fecha}.csv')


def exportar_dia_xlsx(request, fecha_iso: str):
    fecha = interpretar_fecha(fecha_iso)
    consolidado, _ = consolidado_por_fecha(fecha)
    return respuesta_xlsx(consolidado, f'upbase_consolidado_{fecha}.xlsx')


def respuesta_csv(consolidado: list[dict], nombre: str) -> HttpResponse:
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['CODIGO_UP_BASE', 'CANTIDAD', 'PRODUCTO_REFERENCIAL', 'CAJAS'])
    for fila in consolidado:
        writer.writerow([fila['codigo_final'], fila['cantidad'], fila['producto'], fila.get('cajas', '')])
    return response


def respuesta_xlsx(consolidado: list[dict], nombre: str) -> HttpResponse:
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Registro Up Base'
    hoja.append(['Código a digitar en Up Base', 'Cantidad', 'Producto referencial', 'Cajas incluidas'])
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill('solid', fgColor='D9EAF7')
    for fila in consolidado:
        hoja.append([fila['codigo_final'], fila['cantidad'], fila['producto'], fila.get('cajas', '')])
    hoja.column_dimensions['A'].width = 30
    hoja.column_dimensions['B'].width = 14
    hoja.column_dimensions['C'].width = 60
    hoja.column_dimensions['D'].width = 44
    hoja.freeze_panes = 'A2'
    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    response = HttpResponse(
        salida.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response
